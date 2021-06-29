import os
import pdb
import time
import torch
import openpyxl
import numpy as np
import scipy.io as sio
import scipy.sparse as sp
from sklearn.neighbors import NearestNeighbors
from sklearn.preprocessing import minmax_scale, maxabs_scale, normalize, robust_scale, scale


def load_data(direction_path, dataset_name, normalization=True, normalization_type=None, load_saved=True,
              k_nearest_neighobrs=50, prunning_one=False, prunning_two=True, common_neighbors=2):
    # construct entire file path
    file_type = '.' + direction_path.split('/')[-1]
    target_path = direction_path + '/' + dataset_name + file_type
    feature_path = direction_path + '/' + dataset_name + '_X' + file_type

    # load features and lables
    data = sio.loadmat(target_path)
    try:
        features = data['X']  # <class 'numpy.ndarray'> (n_samples, n_fea)
    except KeyError:
        features = sp.load_npz(feature_path).todense()
    labels = data['Y'].flatten()  # <class 'numpy.ndarray'> (n_samples, )
    labels = torch.from_numpy(labels).float()

    if normalization:
        if normalization_type == 'minmax_scale':
            features = minmax_scale(features)
        elif normalization_type == 'maxabs_scale':
            features = maxabs_scale(features)
        elif normalization_type == 'normalize':
            features = normalize(features)
        elif normalization_type == 'robust_scale':
            features = robust_scale(features)
        elif normalization_type == 'scale':
            features = scale(features)
        elif normalization_type == '255':
            features = np.divide(features, 255.)
        elif normalization_type == '50':
            features = np.divide(features, 50.)
        else:
            print("Please enter a correct normalization type!")
            pdb.set_trace()

    save_direction = './data/adj_matrix/' + dataset_name + '/'
    if not os.path.exists(save_direction):
        os.makedirs(save_direction)
    if load_saved:
        print("load the saved adjacency matrix of " + dataset_name)
        adj = sp.load_npz(save_direction + 'adj.npz')
        adj_wave = sp.load_npz(save_direction + 'adj_wave.npz')
        adj_hat = sp.load_npz(save_direction + 'adj_hat.npz')
    else:
        # construct three kinds of adjacency matrix
        print("Constructing the adjacency matrix of " + dataset_name)
        adj, adj_wave, adj_hat = construct_adjacency_matrix(features, k_nearest_neighobrs, prunning_one,
                                                            prunning_two, common_neighbors)
        # save these scale and matrix
        print("Saving the adjacency matrix to " + save_direction)
        sp.save_npz(save_direction + 'adj.npz', adj)
        sp.save_npz(save_direction + 'adj_wave.npz', adj_wave)
        sp.save_npz(save_direction + 'adj_hat.npz', adj_hat)

    # transform to sparse float tensor
    # features = construct_sparse_float_tensor(features)
    features = torch.from_numpy(features).float()
    adj_wave = construct_sparse_float_tensor(adj_wave)
    adj_hat = construct_sparse_float_tensor(adj_hat)

    norm = adj.shape[0] * adj.shape[0] / float((adj.shape[0] * adj.shape[0] - adj.sum()) * 2)  # <class 'float'>
    pos_weight = float(adj.shape[0] * adj.shape[0] - adj.sum()) / adj.sum()  # <class 'numpy.float64'>
    weight_mask = adj_wave.to_dense().view(-1) == 1
    weight_tensor = torch.ones(weight_mask.size(0))
    weight_tensor[weight_mask] = pos_weight  # <class 'torch.Tensor'>

    return features, labels, adj_wave, adj_hat, norm, weight_tensor


def construct_adjacency_matrix(features, k_nearest_neighobrs, prunning_one, prunning_two, common_neighbors):
    start_time = time.time()
    nbrs = NearestNeighbors(n_neighbors=k_nearest_neighobrs+1, algorithm='ball_tree').fit(features)
    adj_wave = nbrs.kneighbors_graph(features)  # <class 'scipy.sparse.csr.csr_matrix'>
    print("The time cost of construction: ", time.time() - start_time)

    if prunning_one:
        # Pruning strategy 1
        original_adj_wave = adj_wave.A
        judges_matrix = original_adj_wave == original_adj_wave.T
        np_adj_wave = original_adj_wave * judges_matrix
        adj_wave = sp.csc_matrix(np_adj_wave)
    else:
        # transform the matrix to be symmetric (Instead of Pruning strategy 1)
        np_adj_wave = construct_symmetric_matrix(adj_wave.A)
        adj_wave = sp.csc_matrix(np_adj_wave)

    # obtain the adjacency matrix without self-connection
    adj = sp.csc_matrix(np_adj_wave)
    adj = adj - sp.dia_matrix((adj.diagonal()[np.newaxis, :], [0]), shape=adj.shape)
    adj.eliminate_zeros()

    if prunning_two:
        # Pruning strategy 2
        adj = adj.A
        b = np.nonzero(adj)
        rows = b[0]
        cols = b[1]
        dic = {}
        for row, col in zip(rows, cols):
            if row in dic.keys():
                dic[row].append(col)
            else:
                dic[row] = []
                dic[row].append(col)
        for row, col in zip(rows, cols):
            if len(set(dic[row]) & set(dic[col])) < common_neighbors:
                adj[row][col] = 0
        adj = sp.csc_matrix(adj)
        adj.eliminate_zeros()

    # construct the adjacency hat matrix
    adj_hat = construct_adjacency_hat(adj)  # <class 'scipy.sparse.coo.coo_matrix'>  (n_samples, n_samples)
    print("The construction of adjacency matrix is finished!")
    print("The time cost of construction: ", time.time() - start_time)

    return adj, adj_wave, adj_hat


def construct_adjacency_hat(adj):
    """
    :param adj: original adjacency matrix  <class 'scipy.sparse.csr.csr_matrix'>
    :return:
    """
    adj = sp.coo_matrix(adj)
    adj_ = adj + sp.eye(adj.shape[0])
    rowsum = np.array(adj_.sum(1))  # <class 'numpy.ndarray'> (n_samples, 1)
    degree_mat_inv_sqrt = sp.diags(np.power(rowsum, -0.5).flatten())
    # <class 'scipy.sparse.coo.coo_matrix'>  (n_samples, n_samples)
    adj_normalized = adj_.dot(degree_mat_inv_sqrt).transpose().dot(degree_mat_inv_sqrt).tocoo()
    return adj_normalized


def construct_symmetric_matrix(original_matrix):
    """
        transform a matrix (n*n) to be symmetric
    :param np_matrix: <class 'numpy.ndarray'>
    :return: result_matrix: <class 'numpy.ndarray'>
    """
    result_matrix = np.zeros(original_matrix.shape, dtype=float)
    num = original_matrix.shape[0]
    for i in range(num):
        for j in range(num):
            if original_matrix[i][j] == 0:
                continue
            elif original_matrix[i][j] == 1:
                result_matrix[i][j] = 1
                result_matrix[j][i] = 1
            else:
                print("The value in the original matrix is illegal!")
                pdb.set_trace()
    assert (result_matrix == result_matrix.T).all() == True

    if ~(np.sum(result_matrix, axis=1) > 1).all():
        print("There existing a outlier!")
        pdb.set_trace()

    return result_matrix


def construct_sparse_float_tensor(np_matrix):
    """
        construct a sparse float tensor according a numpy matrix
    :param np_matrix: <class 'numpy.ndarray'>
    :return: torch.sparse.FloatTensor
    """
    sp_matrix = sp.csc_matrix(np_matrix)
    three_tuple = sparse_to_tuple(sp_matrix)
    sparse_tensor = torch.sparse.FloatTensor(torch.LongTensor(three_tuple[0].T),
                                        torch.FloatTensor(three_tuple[1]),
                                        torch.Size(three_tuple[2]))
    return sparse_tensor


def sparse_to_tuple(sparse_mx):
    if not sp.isspmatrix_coo(sparse_mx):
        sparse_mx = sparse_mx.tocoo()

    # sparse_mx.row/sparse_mx.col  <class 'numpy.ndarray'> [   0    0    0 ... 2687 2694 2706]
    coords = np.vstack((sparse_mx.row, sparse_mx.col)).transpose()  # <class 'numpy.ndarray'> (n_edges, 2)
    values = sparse_mx.data  # <class 'numpy.ndarray'> (n_edges,) [1 1 1 ... 1 1 1]
    shape = sparse_mx.shape  # <class 'tuple'>  (n_samples, n_samples)
    return coords, values, shape


def save_training_process(dataset_name, curr_repeate, List_ID, List_loss, List_ACC, List_NMI, List_ARI):
    if len(List_ID) == 0:
        pass
    save_path = dataset_name + "_" + str(curr_repeate) + "_training_process.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = str(curr_repeate)
    for i in range(len(List_loss)):
        sheet.cell(row=i + 1, column=1, value=str(List_ID[i]))
        sheet.cell(row=i + 1, column=2, value=str(List_loss[i]))
        sheet.cell(row=i + 1, column=3, value=str(List_ACC[i]))
        sheet.cell(row=i + 1, column=4, value=str(List_NMI[i]))
        sheet.cell(row=i + 1, column=5, value=str(List_ARI[i]))
    workbook.save(save_path)
    print("Finished the save of training process！")


if __name__ == '__main__':
    features = np.array([[1,2,3,4], [2,3,4,5], [6,7,8,9], [3,4,5,6], [4,5,6,7], [5,6,7,8]])
    nbrs = NearestNeighbors(n_neighbors=5, algorithm='ball_tree').fit(features)
    adj_wave = nbrs.kneighbors_graph(features)  # <class 'scipy.sparse.csr.csr_matrix'>
    original_adj_wave = adj_wave.A
    print("original adj wave: ")
    print(original_adj_wave)
    print(adj_wave.A.dtype)

    judges_matrix = original_adj_wave == original_adj_wave.T
    print("Pruning strategy 1: ")
    np_adj_wave = original_adj_wave * judges_matrix
    print(np_adj_wave)

    # obtain the adjacency matrix without self-connection
    adj = sp.csc_matrix(np_adj_wave)
    adj = adj - sp.dia_matrix((adj.diagonal()[np.newaxis, :], [0]), shape=adj.shape)
    adj.eliminate_zeros()
    print("Remove self-connection: ")
    adj = adj.A
    print(adj)

    b = np.nonzero(adj)
    print("b: ", b)
    rows = b[0]
    cols = b[1]
    dic = {}
    for row, col in zip(rows, cols):
        if row in dic.keys():
            dic[row].append(col)
        else:
            dic[row] = []
            dic[row].append(col)
    print("dic: ")
    print(dic)

    new_adj = adj.copy()

    for row, col in zip(rows, cols):
        if len(set(dic[row]) & set(dic[col])) < 2:
            new_adj[row][col] = 0

    print("Pruning Strategy 2: ")
    print(new_adj)


